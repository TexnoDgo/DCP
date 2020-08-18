from openpyxl import load_workbook

wb = load_workbook('C:/PP/DCProject/DCP/TestPart/Example.xlsx')

sheet = wb['Лист1']
# Информация о заказе
order_title = sheet.cell(row=2, column=2).value
order_project = sheet.cell(row=3, column=2).value
order_readiness = sheet.cell(row=4, column=2).value

print(order_readiness,
      order_project,
      order_title)
i = 1
while i != 0:
    row_0 = 7
    col_0 = 4
    # Информация о детали
    detail_title = sheet.cell(row=row_0+i, column=1).value
    detail_material = sheet.cell(row=row_0+i, column=3).value
    detail_assortment = sheet.cell(row=row_0+i, column=4).value
    # Информация о позициях
    position_quantity = sheet.cell(row=row_0 + i, column=2).value
    # Информация об операциях
    print(detail_title, position_quantity, detail_material, detail_assortment)
    for a in range(1, 8):
        operation = sheet.cell(row=row_0+i, column=col_0+a).value
        operation_manufactured = sheet.cell(row=7, column=col_0+a).value
        if operation is not None:
            print(operation_manufactured, operation)
    if detail_title is not None:
        i += 1
    else:
        i = 0
