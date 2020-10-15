import os
from openpyxl import load_workbook
import urllib.request

from django.views.generic import ListView, View, DetailView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin

from django.utils.crypto import get_random_string
from django.contrib import messages
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate


from .models import (Assortment, Detail,
                     Project, Order, Position, City,
                     Manufactured, Operation, Transaction, StockageCode, SystemFile, Fields_Position, blog)
from .forms import (ProjectCreateForm, AssortmentCreateForm,
                    DetailCreateForm, OrderCreateForm, OrderSuperCreateForm, OrderDRAWUploadForm,
                    PositionCreateForm, CityCreateForm, ManufacturedCreateForm, OperationCreateForm,
                    TransactionCreateForm, PositionStorageForm, PositionSearch, PositionDrawAdd,
                    SignUpForm, PasitionAddForm, blog_create_form)
from .handlers import convert_pdf_to_bnp, qr_generator, create_pdf, detail_check, ex_archive, pdf_archive_form

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def index(request):
    details = Detail.objects.all()
    manufactureds = Manufactured.objects.all()
    detail_list = []
    for detail in details:
      detail_list.append(detail.title)
    all_pos = Position.objects.filter(detail=None)
    positions = None
    text = ''

    if request.method == 'GET':

        form = PositionSearch(request.GET)
        print('post')

        if form.is_valid():
            print('valid')
            my_request = form.cleaned_data.get('my_request', None)
            try:
                positions = Position.objects.filter(detail__title=my_request)
                print(len(positions))
                if len(positions) == 0:
                    text = 'Данная деталь отсутствует в системе'
            except:
                print('ex')
                text = 'Данная деталь отсутствует в системе'
                messages.success(request,
                                 # Формирование сообщения со вложенным именем
                                 f'Данной детали не существует. ')

    else:
        form = PositionSearch()
    print(text)
    context = {
        'details': details,
        'form': form,
        'all_pos': positions,
        'text': text,
        'detail_list': detail_list,
        'manufactureds': manufactureds,
    }
    return render(request, 'MainApp/HomePage.html', context)


@login_required
def detail_create(request):
    if request.method == "POST":

        form = DetailCreateForm(request.POST, request.FILES)
        print('post')

        if form.is_valid():
            detail = form.save(commit=False)
            detail.author = request.user
            form.save()
            pdf_file_name = str(detail.draw_pdf)
            print(pdf_file_name)  # Delete
            png_file_name = '{}{}'.format(pdf_file_name[9:-3], 'png')
            print(png_file_name)  # Delete
            png_full_path = os.path.join(BASE_DIR, 'media/PNG_COVER/') + png_file_name
            print(png_full_path)  # Delete
            convert_pdf_to_bnp(detail.draw_pdf.path, png_full_path)
            png_path_name = 'PNG_COVER/' + png_file_name
            print(png_path_name)
            detail.draw_png = png_path_name
            detail.save()
            print('valid')
            return redirect('detail_all')
    else:

        form = DetailCreateForm()
        print('else')

    context = {
        'form': form,
    }

    return render(request, 'MainApp/Detail_Create.html', context)


@login_required
def detail_view(request, url):
    detail = Detail.objects.get(pk=url)
    
    context = {
        'detail': detail,
    }
    
    return render(request, 'MainApp/DetailView.html', context)


@login_required
def details_all(request):
    all_detail = Detail.objects.all()
    manufactureds = Manufactured.objects.all()

    context = {
        'all_detail': all_detail,
        'manufactureds': manufactureds,
    }
    return render(request, 'MainApp/All_Details.html', context)


class DetailUpdate(UpdateView):

    model = Detail
    
    def get_context_data(self, **kwargs):
        context = super(DetailUpdate, self).get_context_data(**kwargs)
        # a = self.object.id
        
        return context
    
    fields = ['title', 'draw_pdf', 'material', 'assortment', 'thickness_diameter']
    
    def form_valid(self, form):
        form.instance.author = self.request.user
        detail = form.save(commit=False)
        form.save()
        pdf_file_name = str(detail.draw_pdf)
        png_file_name = '{}{}'.format(pdf_file_name[9:-3], 'png')
        png_full_path = os.path.join(BASE_DIR, 'media/PNG_COVER/') + png_file_name
        convert_pdf_to_bnp(detail.draw_pdf.path, png_full_path)
        png_path_name = 'PNG_COVER/' + png_file_name
        detail.draw_png = png_path_name
        detail.save()
        return redirect('orders_all')


@login_required
def order_create(request):

    if request.method == "POST":

        form = OrderCreateForm(request.POST)

        if form.is_valid():
            order = form.save(commit=False)
            order.author = request.user
            order.save()
            return redirect('orders_all')
    else:
        form = OrderCreateForm()

    context = {
        'form': form,
    }

    return render(request, 'MainApp/Order_Create.html', context)


@login_required
def order_super_create(request):
    if request.method == 'POST':

        order_form = OrderSuperCreateForm(request.POST, request.FILES)

        if order_form.is_valid():
            super_order = order_form.save(commit=False)
            wb = load_workbook(super_order.table)
            sheet = wb['Лист1']
            # ----------------Информация о заказе----------------
            order_title = sheet.cell(row=2, column=2).value
            order_project = sheet.cell(row=3, column=2).value
            order_readiness = sheet.cell(row=4, column=2).value
            order_readiness = (str(order_readiness))[0:10]
            super_order.title = order_title
            super_order.author = request.user
            ex_project = Project.objects.get(title=order_project)
            super_order.project = ex_project
            super_order.readiness = order_readiness
            super_order.save()
            # ----------------Информация о заказе----------------

            # ----------------Информация по позициям----------------
            i = 1
            while i != 0:
                row_0 = 7
                col_0 = 5
                # Информация о детали
                detail_title = sheet.cell(row=row_0 + i, column=1).value
                detail_material = sheet.cell(row=row_0 + i, column=3).value
                detail_assortment = sheet.cell(row=row_0 + i, column=4).value
                detail_thickness_diameter = sheet.cell(row=row_0 + i, column=5).value
                # Информация о позициях
                position_quantity = sheet.cell(row=row_0 + i, column=2).value

                if detail_title is not None:
                    i += 1
                    print(detail_title)
                    ex_part = detail_check(detail_title)
                    print(ex_part)
                    # --------------------Создание детали-----------------------
                    # ex_material = Material.objects.get(title=detail_material)
                    ex_material = detail_material
                    ex_assortment = Assortment.objects.get(title=detail_assortment)
                    ex_stockage_code = StockageCode.objects.get(title='Без расположения')
                    code = get_random_string(length=32)
                    qr_code = qr_generator(code)
                    if not ex_part:
                        # --------------------Создание детали-----------------------
                        detail = Detail(title=detail_title, author=request.user, material=ex_material,
                                        assortment=ex_assortment, thickness_diameter=detail_thickness_diameter)
                        detail.save()
                        # --------------------Создание детали-----------------------
                        # --------------------Создание позиции-----------------------
                        position = Position(order=super_order, detail=detail, quantity=position_quantity,
                                            code=code, qr_code=qr_code, stockage_code=ex_stockage_code)
                        # --------------------Создание позиции-----------------------
                        position.save()
                    elif ex_part:
                        detail = Detail.objects.get(title=detail_title)
                        position = Position(order=super_order, detail=detail, quantity=position_quantity,
                                            code=code, qr_code=qr_code, stockage_code=ex_stockage_code)
                        position.save()
                    else:
                        print('IF error')

                   

                    # ----------------------Информация об операциях-----------------
                    for a in range(1, 8):
                        operation = sheet.cell(row=row_0 + i - 1, column=col_0 + a).value
                        operation_manufactured = sheet.cell(row=7, column=col_0 + a).value
                        if operation is not None:
                            ex_manufactured = Manufactured.objects.get(title=operation_manufactured)
                            ex_opertion = Operation(manufactured=ex_manufactured, position=position,
                                                    remaining_parts=position_quantity)
                            ex_opertion.save()
                            '''if ex_opertion.manufactured.title == 'Опытынй завод':
                                fp.operation_oz = ex_opertion
                            elif ex_opertion.manufactured.title == 'НИИ Лазерная Резка':
                                fp.operation_niilr = ex_opertion
                            elif ex_opertion.manufactured.title == 'Альянс Сталь':
                                fp.operation_alianse = ex_opertion
                            elif ex_opertion.manufactured.title == 'CNC MetalWork':
                                fp.operation_cncmw = ex_opertion
                            elif ex_opertion.manufactured.title == 'Покрытие №1':
                                fp.operation_pk1 = ex_opertion
                            elif ex_opertion.manufactured.title == 'Покрытие №2':
                                fp.operation_pk2 = ex_opertion
                            elif ex_opertion.manufactured.title == 'Другой':
                                fp.operation_dr = ex_opertion'''
                            if ex_opertion.manufactured.title == 'Опытынй завод':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_oz = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_oz=ex_opertion)
                                    fp.save()
                            elif ex_opertion.manufactured.title == 'НИИ Лазерная Резка':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_niilr = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_niilr=ex_opertion)
                                    fp.save()
                            elif ex_opertion.manufactured.title == 'Альянс Сталь':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_alianse = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_alianse=ex_opertion)
                                    fp.save()
                            elif ex_opertion.manufactured.title == 'CNC MetalWork':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_cncmw = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_cncmw=ex_opertion)
                                    fp.save()
                            elif ex_opertion.manufactured.title == 'Покрытие №1':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_pk1 = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_pk1=ex_opertion)
                                    fp.save()
                            elif ex_opertion.manufactured.title == 'Покрытие №2':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_pk2 = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_pk2=ex_opertion)
                                    fp.save()
                            elif ex_opertion.manufactured.title == 'Другой':
                                try:
                                    fp = Fields_Position.objects.get(position=position)
                                    fp.operation_dr = ex_opertion
                                    fp.save()
                                except:
                                    fp = Fields_Position(position=position, operation_dr=ex_opertion)
                                    fp.save()
                            fp.save()
                            
                    # ----------------------Информация об операциях-----------------
                else:
                    i = 0
            # ----------------Информация по позициям----------------
            pdf_file_path = create_pdf(super_order)
            super_order.qr_code_list = pdf_file_path
            super_order.save()
            return redirect('orders_all')
    else:
        order_form = OrderSuperCreateForm()

    context = {
        'order_form': order_form,
    }

    return render(request, 'MainApp/Order_Super_Create.html', context)


@login_required
def orders_all(request):

    all_orders = Order.objects.all().order_by('-pk')
    manufactureds = Manufactured.objects.all()
    positions = Position.objects.all()

    '''orders = {}
    for order in all_orders:
        orders = order.title
        positions = Position.objects.filter(order=order)
        position_quantity = 0
        for position in positions:
            position_quantity += position.quantity
            operations = Operation.objects.filter(position=position)
            operation_remaining_parts = 0
            for operation in operations:
                operation_remaining_parts += int(operation.remaining_parts)
            orders[order.title]['rem'] = operation_remaining_parts
            orders[order.title]['qua'] = position_quantity
        print(orders)'''

    context = {
        'all_orders': all_orders,
        'manufactureds': manufactureds,
        'positions': positions,
    }

    return render(request, 'MainApp/All_Orders.html', context)


@login_required
def order_view(request, url):
    order = Order.objects.get(pk=url)
    positions = Position.objects.filter(order=order)
    details = Detail.objects.all()
    operations = Operation.objects.all()
    manufactureds = Manufactured.objects.all()

    ready_positions = {}

    for position in positions:
        order_operations = Operation.objects.filter(position=position)
        ready_operations = 0
        for operation in order_operations:
            ready_operations += int(operation.remaining_parts)
        ready_positions[position.detail.title] = ready_operations

    print(ready_positions)

    if request.method == 'POST':

        form = OrderDRAWUploadForm(request.POST, request.FILES)

        if form.is_valid():
            archive = form.cleaned_data.get('archive', None)
            flag = form.cleaned_data.get('flag', None)
            order.draw_archive = archive
            order.save()
            if flag:
                print(ex_archive(order))
            return redirect(request.META['HTTP_REFERER'])
    else:

        form = OrderDRAWUploadForm()

    context = {
        'order': order,
        'positions': positions,
        'details': details,
        'operations': operations,
        'form': form,
        'ready_positions': ready_positions,
        'manufactureds': manufactureds,
    }

    return render(request, 'MainApp/Order.html', context)


@login_required
def add_pisition(request, url):
    order = Order.objects.get(pk=url)
    
    if request.method == 'POST':
    
        form = PasitionAddForm(request.POST, request.FILES)
        
        if form.is_valid():
            detail_title = form.cleaned_data.get('detail_title', None)
            quantity = form.cleaned_data.get('quantity', None)
            #draw_pdf = form.cleaned_data.get('draw_pdf', None)
            material = form.cleaned_data.get('material', None)
            assortment = form.cleaned_data.get('assortment', None)
            thickness_diameter = form.cleaned_data.get('thickness_diameter', None)
            
            ex_part = detail_check(detail_title)
            code = get_random_string(length=32)
            qr_code = qr_generator(code)
            ex_stockage_code = StockageCode.objects.get(title='Без расположения')
            try:
                ex_assortmaent = Assortment.objects.get(title=assortment)
            except:
                messages.success(request, f'ASSORTMENT ERROR!')
            
            if not ex_part:
            
                detail = Detail(title=detail_title, author=request.user, material=material,
                                assortment=ex_assortmaent, thickness_diameter=thickness_diameter)
                detail.save()
                
                position = Position(order=order, detail=detail, quantity=quantity,
                                    code=code, qr_code=qr_code, stockage_code=ex_stockage_code)

            elif ex_part:
                detail = Detail.objects.get(title=detail_title)
                
                position = Position(order=order, detail=detail, quantity=quantity,
                                    code=code, qr_code=qr_code, stockage_code=ex_stockage_code)
            else:
                print('IF error')

            position.save()
            
            fp = Fields_Position(position=position)
            fp.save()

    else:
        form = PasitionAddForm()
        messages.success(request, f'FORM DIDNT VALID!')
        
    
    context = {
        'order': order,
        'form': form,
    }
    return render(request, 'MainApp/PositionADD.html', context)


@login_required
def position_view(request, code):
    position = Position.objects.get(code=code)
    operations = Operation.objects.filter(position=position.pk)
    order = Order.objects.get(pk=position.order.pk)
    detail = Detail.objects.get(pk=position.detail.pk)
    maps = SystemFile.objects.get(title='maps')
    stockage = StockageCode.objects.all()
    manufactureds = Manufactured.objects.all()

    if request.method == 'POST':

        form = PositionStorageForm(request.POST)

        if form.is_valid():
            position_in = form.cleaned_data.get('position_in', None)
            print(position_in)
            try:
                stockage = StockageCode.objects.get(title=position_in)
                position.stockage_code = stockage
                position.save()
            except:
                messages.success(request,
                                 # Формирование сообщения со вложенным именем
                                 f'Данного места не существует. ')

            return redirect(request.META['HTTP_REFERER'])
    else:
        form = PositionStorageForm()

    context = {
        'position': position,
        'operations': operations,
        'order': order,
        'detail': detail,
        'form': form,
        'maps': maps,
        'stockage': stockage,
        'manufactureds': manufactureds,
    }

    return render(request, 'MainApp/Position.html', context)


class PositionUpdate(UpdateView):

    model = Position
    
    
    def get_context_data(self, **kwargs):
        context = super(PositionUpdate, self).get_context_data(**kwargs)
        # a = self.object.id
        
        return context
    
    fields = ['order', 'detail', 'quantity', 'stockage_code']
    
    def form_valid(self, form):
        position = form.save(commit=False)
        position.save()
        code = position.code
        return redirect('position_view', code)


@login_required
def position_draw_change(request, code):
    position = Position.objects.get(code=code)
    detail = Detail.objects.get(pk=position.detail.pk)
    
    if request.method == 'POST':
        print('post')
        
        form = PositionDrawAdd(request.POST, request.FILES)
        
        if form.is_valid():
            draw = form.cleaned_data.get('draw', None)
            detail.draw_pdf = draw
            detail.save()
            pdf_file_name = str(detail.draw_pdf)
            print(pdf_file_name)  # Delete
            png_file_name = '{}{}'.format(pdf_file_name[9:-3], 'png')
            print(png_file_name)  # Delete
            png_full_path = os.path.join(BASE_DIR, 'media/PNG_COVER/') + png_file_name
            print(png_full_path)  # Delete
            convert_pdf_to_bnp(detail.draw_pdf.path, png_full_path)
            png_path_name = 'PNG_COVER/' + png_file_name
            print(png_path_name)
            detail.draw_png = png_path_name
            detail.save()
            
            return redirect('position_view', code)
    
    else:
        print('else')
        form = PositionDrawAdd()
    
    context = {
      'position': position,
      'detail': detail,
      'form': form,
    }
    
    return render(request, 'MainApp/PositionDrawAdd.html', context)
        


@login_required
def operation_view(request, url):
    operation = Operation.objects.get(pk=url)
    transactions = Transaction.objects.filter(operation=operation)

    difference = int(operation.position.quantity) - int(operation.remaining_parts)
    
    manufactureds = Manufactured.objects.all()

    if request.method == 'POST':

        form = TransactionCreateForm(request.POST)

        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.author = request.user
            transaction.operation = operation
            print(operation.remaining_parts)
            if transaction.ready_quantity <= operation.remaining_parts:
                operation.remaining_parts -= transaction.ready_quantity
                print(operation.remaining_parts)
                operation.status = 'PD'
                if operation.remaining_parts == 0:
                    operation.status = 'RD'
                transaction.save()
                operation.save()
            else:
                messages.success(request,
                                 # Формирование сообщения со вложенным именем
                                 f'Количество выполненных деталей больше необходимого. '
                                 f'Пожалуйста проверьте вводимое значение или обратитесь '
                                 f'к автору заказа! ')

            return redirect(request.META['HTTP_REFERER'])
    else:

        form = TransactionCreateForm()

    context = {
        'operation': operation,
        'form': form,
        'difference': difference,
        'transactions': transactions,
        'manufactureds': manufactureds,
    }
    return render(request, 'MainApp/Operation.html', context)


def operation_add(request, code):
    position = Position.objects.get(code=code)
    fp = Fields_Position.objects.get(position=position)
    
    if request.method == 'POST':
        form = OperationCreateForm(request.POST)
        print('post')
        
        if form.is_valid():
            print('valid')
            operation = form.save(commit=False)
            operation.title = "DEFAULT"
            operation.status = "CD"
            operation.position = position
            operation.remaining_parts = position.quantity
            operation.save()
            
            if operation.manufactured.title == 'Опытынй завод':
                fp.operation_oz = operation
            elif operation.manufactured.title == 'НИИ Лазерная Резка':
                fp.operation_niilr = operation
            elif operation.manufactured.title == 'Альянс Сталь':
                fp.operation_alianse = operation
            elif operation.manufactured.title == 'CNC MetalWork':
                fp.operation_cncmw = operation
            elif operation.manufactured.title == 'Покрытие №1':
                fp.operation_pk1 = operation
            elif operation.manufactured.title == 'Покрытие №2':
                fp.operation_pk2 = operation
            elif operation.manufactured.title == 'Другой':
                fp.operation_dr = operation
            fp.save()
            
            messages.success(request, f'Operation ADD!')
            return redirect('position_view', code)
        
    else:
        form = OperationCreateForm()
        print('else')
        
    context = {
        'form': form,
        'position': position,
    }
    
    return render(request, 'MainApp/OperationADD.html', context)


class OperationDelete(DeleteView):

    model = Operation

    success_url = '/all/AllOrders'

    def test_func(self):
        operation = self.get_object()
        if self.request.user == operation.position.order.author:
            return True
        return False

    def get_context_data(self, **kwargs):
        context = super(OperationDelete, self).get_context_data(**kwargs)
        a = self.object.id
        operation = self.get_object()
        position = Position.objects.get(operation=operation)
        fp = Fields_Position.objects.get(position=position) 
        
        if operation.manufactured.title == 'Опытынй завод':
            fp.operation_oz = None
        elif operation.manufactured.title == 'НИИ Лазерная Резка':
            fp.operation_niilr = None
        elif operation.manufactured.title == 'Альянс Сталь':
            fp.operation_alianse = None
        elif operation.manufactured.title == 'CNC MetalWork':
            fp.operation_cncmw = None
        elif operation.manufactured.title == 'Покрытие №1':
            fp.operation_pk1 = None
        elif operation.manufactured.title == 'Покрытие №2':
            fp.operation_pk2 = None
        elif operation.manufactured.title == 'Другой':
            fp.operation_dr = None

        fp.save()
        
        context['detail'] = operation.position.detail
        return context


def crete_order_qr_code_list(request, url):
    order = Order.objects.get(pk=url)
    create_pdf(order)
    return redirect(request.META['HTTP_REFERER'])


def archive_pdf_former(request, url):
    order = Order.objects.get(pk=url)
    if order.draw_archive:
        pdf_archive_form(url)
    else:
        messages.success(request,
                         # Формирование сообщения со вложенным именем
                         f'Архив с чертежами отсутствует!.')
    # ADD MESSAGE
    return redirect(request.META['HTTP_REFERER'])

@login_required
def manufactured_acc(request, url):
    manufactureds = Manufactured.objects.all()
    manufactured = Manufactured.objects.get(pk=url)
    operations = Operation.objects.filter(manufactured=manufactured).order_by('-pk')
    positions = Position.objects.all().order_by('-pk')

    context = {
        'manufactureds': manufactureds,
        'manufactured': manufactured,
        'operations': operations,
        'positions': positions,
    }

    return render(request, 'MainApp/Manufactured.html', context)
    
    
@login_required
def in_made_status(request, url):
    operation = Operation.objects.get(pk=url)
    if operation.status == 'CD':
        operation.status = 'PD'
    elif operation.status == 'PD':
        operation.status = 'CD'
    operation.save()
    return redirect(request.META['HTTP_REFERER'])


@login_required
def ready_status(request, url):
    operation = Operation.objects.get(pk=url)
    if operation.status == 'PD':
        operation.status = 'RD'
    elif operation.status == 'RD':
        operation.status = 'PD'
    operation.save()
    return redirect(request.META['HTTP_REFERER'])

@login_required
def order_operation_change_status(request, url):
    operation = Operation.objects.get(pk=url)
    if operation.status == 'CD':
        operation.status = 'PD'
    elif operation.status == 'PD':
        operation.status = 'RD'
        operation.remaining_parts = 0
    elif operation.status == 'RD':
        operation.status = 'PD'
        operation.remaining_parts = operation.position.quantity
    operation.save()
    return redirect(request.META['HTTP_REFERER'])

def signup(request):
    if request.method == 'POST':
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            user.refresh_from_db()  # load the profile instance created by the signal
            user.profile.birth_date = form.cleaned_data.get('birth_date')
            user.save()
            raw_password = form.cleaned_data.get('password1')
            user = authenticate(username=user.username, password=raw_password)
            login(request, user)
            return redirect('HomePage')
    else:
        form = SignUpForm()
    return render(request, 'MainApp/signup.html', {'form': form})
    

def bambardier(request):
    positions = Position.objects.all()
    operations = Operation.objects.all()
    for position in positions:
        for operation in operations:
            if operation.position == position:
                if operation.manufactured.title == 'Опытынй завод':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_oz = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_oz=operation)
                        fp.save()
                elif operation.manufactured.title == 'НИИ Лазерная Резка':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_niilr = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_niilr=operation)
                        fp.save()
                elif operation.manufactured.title == 'Альянс Сталь':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_alianse = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_alianse=operation)
                        fp.save()
                elif operation.manufactured.title == 'CNC MetalWork':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_cncmw = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_cncmw=operation)
                        fp.save()
                elif operation.manufactured.title == 'Покрытие №1':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_pk1 = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_pk1=operation)
                        fp.save()
                elif operation.manufactured.title == 'Покрытие №2':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_pk2 = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_pk2=operation)
                        fp.save()
                elif operation.manufactured.title == 'Другой':
                    try:
                        fp = Fields_Position.objects.get(position=position)
                        fp.operation_dr = operation
                        fp.save()
                    except:
                        fp = Fields_Position(position=position, operation_dr=operation)
                        fp.save()
    return redirect(request.META['HTTP_REFERER'])


def fp_view(request, url):
    order = Order.objects.get(pk=url)
    fps = Fields_Position.objects.filter(position__order=order)
    ready_positions = {}
    
    for fp in fps:
        ready_operations = 0

        if fp.operation_oz:
            ready_operations += int(fp.operation_oz.remaining_parts)
        if fp.operation_niilr:
            ready_operations += int(fp.operation_niilr.remaining_parts)
        if fp.operation_alianse:
            ready_operations += int(fp.operation_alianse.remaining_parts)
        if fp.operation_cncmw:
            ready_operations += int(fp.operation_cncmw.remaining_parts)
        if fp.operation_pk1:
            ready_operations += int(fp.operation_pk1.remaining_parts)
        if fp.operation_pk2:
            ready_operations += int(fp.operation_pk2.remaining_parts)
        if fp.operation_dr:
            ready_operations += int(fp.operation_dr.remaining_parts)

        ready_positions[fp.position.detail.title] = ready_operations
    
    if request.method == 'POST':

        form = OrderDRAWUploadForm(request.POST, request.FILES)

        if form.is_valid():
            archive = form.cleaned_data.get('archive', None)
            flag = form.cleaned_data.get('flag', None)
            order.draw_archive = archive
            order.save()
            if flag:
                print(ex_archive(order))
            return redirect(request.META['HTTP_REFERER'])
    else:

        form = OrderDRAWUploadForm()
    
    context = {
    
        'fps': fps,
        'order': order,
        'ready_positions': ready_positions,
        'form': form,
        
    }
    
    return render(request, 'MainApp/FP_VIEW.html', context)


@login_required
def blog_create(request):
    
    articles = blog.objects.all().order_by('create')
    
    if request.method == 'POST':
    
        form = blog_create_form(request.POST)
        messages.success(request, f'post!')
        if form.is_valid():
        
            atr = form.save(commit=False)
            atr.author = request.user
            form.save()
            messages.success(request, f'valid!')
            return redirect(request.META['HTTP_REFERER'])
    else:
        messages.success(request, f'else!')
        form = blog_create_form()
    
    context = {
        'form': form,
        'articles': articles,
    }
    return render(request, 'MainApp/blog.html', context)


def blog_delete(request, url):
    pass

